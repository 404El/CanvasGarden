import os
import json
import requests
import pandas as pd
import re 
from datetime import datetime
from dotenv import load_dotenv
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- PORTABLE FILE PATHS ---
current_dir = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(current_dir)
ENV_PATH = os.path.join(BASE_DIR, 'ugly stuff', '.env')
SETTINGS_FILE = os.path.join(BASE_DIR, 'ugly stuff', 'course_settings.json')

load_dotenv(dotenv_path=ENV_PATH)

CANVAS_BASE_URL = os.getenv("CANVAS_BASE_URL")
CANVAS_ACCESS_TOKEN = os.getenv("CANVAS_ACCESS_TOKEN")

def clean_html(raw_html):
    if not raw_html: return ""
    cleanr = re.compile('<.*?>')
    cleantext = re.sub(cleanr, '', raw_html)
    return cleantext.strip()

def get_category(name):
    name_lower = name.lower()
    if "final" in name_lower: return "Final"
    elif "midterm" in name_lower: return "Midterm"
    elif "test" in name_lower or "exam" in name_lower: return "Test"
    elif "quiz" in name_lower: return "Quiz"
    elif "lab" in name_lower: return "Lab"
    elif "project" in name_lower: return "Project"
    elif any(word in name_lower for word in ["hw", "homework", "assignment"]): return "Homework"
    else: return "Other"

def get_all_assignments():
    if not CANVAS_BASE_URL or not CANVAS_ACCESS_TOKEN:
        print(f"❌ Missing API keys in: {ENV_PATH}")
        return
    if not os.path.exists(SETTINGS_FILE):
        print(f"❌ Missing settings in: {SETTINGS_FILE}")
        return
    
    with open(SETTINGS_FILE, 'r') as f:
        settings_data = json.load(f)
    
    course_settings = settings_data.get("courses", {})
    color_map = {info['name']: info['color'].replace('#', '') for cid, info in course_settings.items()}
    
    all_data = []
    headers = {"Authorization": f"Bearer {CANVAS_ACCESS_TOKEN}"}
    now = datetime.now()

    for cid, info in course_settings.items():
        url = f"{CANVAS_BASE_URL.rstrip('/')}/api/v1/courses/{cid.strip()}/assignments"
        try:
            response = requests.get(url, headers=headers)
            assignments = response.json()
            
            for a in assignments:
                due_date_raw = a.get("due_at")
                due_date = pd.to_datetime(due_date_raw) if due_date_raw else None
                days_left = (due_date.replace(tzinfo=None) - now).days if due_date else 999

                all_data.append({
                    "Course": info.get("name"),
                    "Category": get_category(a.get("name")),
                    "Assignment": a.get("name"),
                    "Due Date": due_date, 
                    "Days Left": days_left,
                    "Points": a.get("points_possible", 0),
                    "Description": clean_html(a.get("description", "")), 
                    "Link": a.get("html_url", "")
                })
        except: continue

    if not all_data:
        print("🍃 No assignments found to harvest.")
        return

    df = pd.DataFrame(all_data).sort_values(by="Days Left", ascending=True)
    course_names_in_order = df['Course'].tolist()

    df['Due Date'] = df['Due Date'].dt.strftime('%b %d').fillna("No Date")
    df['Days Left'] = df['Days Left'].apply(lambda x: x if x != 999 else "N/A")

    filename = os.path.join(BASE_DIR, "Canvas_Assignments.xlsx")
    
    try:
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Assignments')
        ws = writer.sheets['Assignments']

        # --- Earthy Styling Tools ---
        header_fill = PatternFill(start_color='B2AC88', end_color='B2AC88', fill_type='solid')
        header_font = Font(name='Georgia', bold=True, color="FFFFFF", size=11)
        body_font = Font(name='Verdana', size=10, color="4B3621") # Dark Espresso text
        
        # Defining a darker espresso border
        # Changed style to 'thin' but with a much darker color for better visibility
        # If you want even thicker lines, change 'thin' to 'medium'
        dark_side = Side(style='thin', color='3E2723') 
        dark_border = Border(left=dark_side, right=dark_side, top=dark_side, bottom=dark_side)

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
            if row_idx == 1:
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = dark_border
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            else:
                current_course_name = course_names_in_order[row_idx - 2]
                hex_color = color_map.get(current_course_name, 'FDF5E6') 
                course_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

                for cell in row:
                    cell.font = body_font
                    cell.fill = course_fill
                    cell.border = dark_border
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

        # Formatting Column Widths
        ws.column_dimensions['A'].width = 22 # Course
        ws.column_dimensions['B'].width = 15 # Category
        ws.column_dimensions['C'].width = 38 # Assignment
        ws.column_dimensions['D'].width = 12 # Due Date
        ws.column_dimensions['E'].width = 12 # Days Left
        ws.column_dimensions['F'].width = 8  # Points
        ws.column_dimensions['G'].width = 65 # Description
        ws.column_dimensions['H'].width = 40 # Link
        
        writer.close()
        print(f"✨ SUCCESS: Harvest ready at {filename}")
    except Exception as e: print(f"Error: {e}")

if __name__ == "__main__":
    get_all_assignments()